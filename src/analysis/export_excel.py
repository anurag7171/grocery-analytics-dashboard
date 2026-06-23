"""
Build a formatted multi-sheet Excel report from the analysis outputs.

Produces docs/grocery-analytics-report.xlsx with a summary sheet, category and
brand breakdowns, weekly trends, the statistical test results, and the full
product snapshot — styled headers, frozen panes, number formats, and a colour
scale so it reads like an analyst's deliverable, not a raw dump.

Usage (from repo root):
    python -m src.analysis.export_excel
"""

import os

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

BASE       = os.path.join(os.path.dirname(__file__), "..", "..")
TABLEAU    = os.path.join(BASE, "data", "tableau")
RESULTS    = os.path.join(BASE, "results")
OUT_PATH   = os.path.join(BASE, "docs", "grocery-analytics-report.xlsx")

HEADER_FILL = PatternFill("solid", fgColor="16A34A")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT  = Font(bold=True, size=14, color="0F172A")
THIN        = Side(style="thin", color="E2E8F0")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _read(path):
    return pd.read_csv(path) if os.path.exists(path) else pd.DataFrame()


def _style_table(ws, df, currency_cols=(), pct_cols=(), colorscale_cols=()):
    """Apply header style, borders, number formats, widths and freeze panes."""
    ncols = df.shape[1]
    # header row
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    ws.freeze_panes = "A2"

    # column widths + number formats
    for c, col_name in enumerate(df.columns, start=1):
        letter = get_column_letter(c)
        width = max(len(str(col_name)), *(df[col_name].astype(str).map(len).tolist() or [0])) + 3
        ws.column_dimensions[letter].width = min(width, 45)
        fmt = None
        if col_name in currency_cols:
            fmt = '"$"#,##0.00'
        elif col_name in pct_cols:
            fmt = '0.0"%"'
        if fmt:
            for r in range(2, df.shape[0] + 2):
                ws.cell(row=r, column=c).number_format = fmt

    # colour scale on chosen numeric columns
    for col_name in colorscale_cols:
        if col_name in df.columns:
            c = list(df.columns).index(col_name) + 1
            letter = get_column_letter(c)
            rng = f"{letter}2:{letter}{df.shape[0] + 1}"
            ws.conditional_formatting.add(rng, ColorScaleRule(
                start_type="min", start_color="FFFFFF",
                end_type="max", end_color="16A34A",
            ))


def build(out_path: str = OUT_PATH) -> str:
    kpi      = _read(f"{TABLEAU}/category_kpi.csv")
    brand    = _read(f"{TABLEAU}/brand_comparison.csv")
    weekly   = _read(f"{TABLEAU}/weekly_category_avg.csv")
    products = _read(f"{TABLEAU}/products_snapshot.csv")
    ttest    = _read(f"{RESULTS}/brand_ttest.csv")
    inflation = _read(f"{RESULTS}/category_inflation.csv")

    # ── Summary sheet content ────────────────────────────────────────────────
    total_products = int(products.shape[0]) if not products.empty else 0
    avg_price = round(products["price"].mean(), 2) if not products.empty else 0
    on_special = int(products["is_on_special"].sum()) if not products.empty else 0
    summary = pd.DataFrame({
        "Metric": [
            "Products analysed", "Categories", "Average price",
            "Products on special", "Most expensive category",
            "Highest promo activity", "Report generated",
        ],
        "Value": [
            f"{total_products:,}",
            int(kpi.shape[0]) if not kpi.empty else 0,
            f"${avg_price:,.2f}",
            f"{on_special:,} ({round(100*on_special/total_products,1) if total_products else 0}%)",
            kpi.loc[kpi['avg_price'].idxmax(), 'category'] if not kpi.empty else "-",
            kpi.loc[kpi['special_pct'].idxmax(), 'category'] if not kpi.empty else "-",
            pd.Timestamp.utcnow().strftime("%d %b %Y"),
        ],
    })

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as xl:
        summary.to_excel(xl, sheet_name="Summary", index=False, startrow=2)
        kpi.to_excel(xl, sheet_name="Category KPIs", index=False)
        brand.to_excel(xl, sheet_name="Brand Comparison", index=False)
        weekly.to_excel(xl, sheet_name="Weekly Trends", index=False)
        if not ttest.empty:    ttest.to_excel(xl, sheet_name="Brand t-test", index=False)
        if not inflation.empty: inflation.to_excel(xl, sheet_name="Inflation", index=False)
        products.to_excel(xl, sheet_name="Products", index=False)

        # Title on the summary sheet
        ws_sum = xl.sheets["Summary"]
        ws_sum["A1"] = "Australian Grocery Price Analytics — Summary"
        ws_sum["A1"].font = TITLE_FONT
        _style_table(ws_sum, summary)
        # the summary header is on row 3 (startrow=2); restyle that header row
        for c in range(1, summary.shape[1] + 1):
            hc = ws_sum.cell(row=3, column=c)
            hc.fill = HEADER_FILL; hc.font = HEADER_FONT
            hc.alignment = Alignment(horizontal="center")
        ws_sum.column_dimensions["A"].width = 28
        ws_sum.column_dimensions["B"].width = 30

        _style_table(xl.sheets["Category KPIs"], kpi,
                     currency_cols=("avg_price", "min_price", "max_price"),
                     pct_cols=("special_pct", "avg_discount_pct"),
                     colorscale_cols=("avg_price",))
        _style_table(xl.sheets["Brand Comparison"], brand,
                     currency_cols=("avg_price",), pct_cols=("avg_discount", "special_pct"))
        _style_table(xl.sheets["Weekly Trends"], weekly,
                     currency_cols=("avg_price", "avg_regular_price"), pct_cols=("special_pct",))
        if not ttest.empty:     _style_table(xl.sheets["Brand t-test"], ttest)
        if not inflation.empty: _style_table(xl.sheets["Inflation"], inflation)
        _style_table(xl.sheets["Products"], products,
                     currency_cols=("price", "was_price"), pct_cols=("discount_pct",))

    print(f"Excel report saved: {out_path}")
    return out_path


if __name__ == "__main__":
    build()
